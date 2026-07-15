import sys, json, argparse
from pathlib import Path
from typing import List, Tuple
import openpyxl
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
from models.emissions_model import get_emissions_by_product
from models.products_model import fetch_product
from models.factor_model import get_factor
from flask import current_app
from routes.helpers import to_taipei_iso

TARGET_STAGE_IN_EXCEL = {
    "原料取得": {
        "anchor": "B33",
        "cols": ["B", "C", "D"]
    },
    "製造": {
        "anchor": "B75",
        "cols": ["B", "C", "D"]
    },
    "配銷":{
        "anchor": "B158",
        "cols": ["B", "C", "D"]
    },
    "使用":{
        "anchor": "B172",
        "cols": ["B", "C", "D"]
    },
    "廢棄":{
        "anchor": "B185",
        "cols": ["B", "C", "D"]
    },
}

PROTECTED_MERGED_RANGES = [
    "B71:M71",
    "B157:O157",
    "B170:M170",
    "B184:M184",
]

WHITE_FILL = PatternFill(fill_type=None)
BLUE_FONT = Font(color="0000FF")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THICK_LEFT = Border(
    left = Side(style="thick"),
    right = Side(style="thin"),
    top = Side(style="thin"),
    bottom= Side(style="thin")
)
THIN = Border(
    left = Side(style="thin"),
    right = Side(style="thin"),
    top = Side(style="thin"),
    bottom= Side(style="thin")
)

STAGE_ENG_TO_ZH = {
    "raw": "原料取得",
    "manufacture": "製造",
    "distribution": "配銷",
    "use": "使用",
    "disposal": "廢棄",
}


STAGE_ORDER_KEYS = ["raw", "manufacture", "distribution", "use", "disposal"]


def generate_json(product_id: int):

    pd = fetch_product(product_id) 
    ems = get_emissions_by_product(product_id) # list of dicts
    
    records_dir = Path(current_app.config["REPORT_RECORDS_DIR"])
    records_dir.mkdir(parents=True, exist_ok=True)

    output_json = f"{product_id}.json"
    output_path = records_dir / output_json

    # structure the data according to the expected JSON format
    data = {
        "product": {
            "id": product_id,
            "name": pd.get('name'),
            "type_id": pd.get('type_id'),
        },
        "stages": []
    }
    
    # stage_key -> list of records
    stage_dict: dict[str, list[dict]] = {}

    for em in ems:
        stage_id = em.get("stage_id")
        if stage_id not in stage_dict:
            stage_dict[stage_id] = []

        record = {
            "material": em.get("name"),
            "factor_id": em.get("factor_id"),
            "amount": em.get("quantity"),
            "unit": (get_factor(em.get("factor_id")) or {}).get("unit", ""), # from factor
            "emission_amount": em.get("emission_amount"),
            "timestamp": to_taipei_iso(em.get("created_at")),
        }
        
        stage_dict[stage_id].append(record)

    for stage_id, records in stage_dict.items():
        data["stages"].append({
            "stage_name": STAGE_ENG_TO_ZH.get(stage_id),
            "records": records
        })
        
   
    # write to output JSON file, under report/records/
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    return output_path


def load_json(p: Path):
    return json.loads(p.read_text(encoding="utf-8"))

def collect_product_name(data):
    product_name = data.get("product", {}).get("name")
    return product_name

def collect_by_stage(data):
    out = {}
    for st in data.get("stages", []):
        mapped = data.get(st.get("stage_name"), st.get("stage_name"))
        rows = []
        for r in st.get("records", []):
            name = r.get("material") or r.get("name") or ""
            amt  = r.get("emission_amount")
            if amt is not None:
                amt = round(amt, 5)
            unit = r.get("unit") or ""
            rows.append((name, amt, unit))
        if rows:
            out.setdefault(mapped, []).extend(rows)
    return out

def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    num = 0
    for ch in letter:
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def _col_letters_to_indices(cols_letters: list) -> list:
    return [col_letter_to_index(x) for x in cols_letters]

def anchor_to_row(anchor: str) -> int:
    # B33 -> 33
    return int(anchor[1:])

def copy_row_style_borders(ws, src_row: int, dst_row: int, cols_letters: list):
    cols = _col_letters_to_indices(cols_letters)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in cols:
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        try: dst.border = copy(src.border)
        except: pass
        dst.number_format = src.number_format
        try: dst.alignment = copy(src.alignment)
        except: pass
        try: dst.font = copy(src.font)
        except: pass

def insert_rows_preserve_merges(ws, idx: int, amount: int = 1):
    if amount <= 0:
        return

    old_ranges = list(ws.merged_cells.ranges)
    merged_info = []
    for mr in old_ranges:
        merged_info.append((mr.min_col, mr.min_row, mr.max_col, mr.max_row))

    for mr in old_ranges:
        ws.unmerge_cells(str(mr))

    ws.insert_rows(idx, amount)
        
    for min_col, min_row, max_col, max_row in merged_info:
        if min_row >= idx:
            min_row += amount
            max_row += amount

        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )

def ensure_space_for_rows(ws, start_row: int, cols_letters: List[str], needed_rows: int) -> int:
    cols = [col_letter_to_index(x) for x in cols_letters]
    max_r = ws.max_row

    r = start_row
    while r <= max_r:
        vals = [ws.cell(row=r, column=c).value for c in cols]
        if all(v in (None, "") for v in vals):
            break
        r += 1
    first_free = r  # first empty row after existing data

    boundary = first_free
    while boundary <= max_r:
        vals = [ws.cell(row=boundary, column=c).value for c in cols]
        if any(v not in (None, "") for v in vals):
            break
        boundary += 1

    capacity = boundary - first_free  # current continuous empty rows
    extra = max(0, needed_rows - capacity)

    if extra > 0:
        insert_rows_preserve_merges(ws, boundary, extra)

    return first_free, extra, boundary

def unmerge_only_on_row_for_cols(ws, row: int, cols_letters: list):
    """
    only unmerge merged cells that overlap with the specified row and column letters.
    """
    target_cols = [col_letter_to_index(c) for c in cols_letters]
    min_col, max_col = min(target_cols), max(target_cols)

    protected_bounds = []
    for coord in PROTECTED_MERGED_RANGES:
        p_min_col, p_min_row, p_max_col, p_max_row = range_boundaries(coord)
        protected_bounds.append((p_min_col, p_min_row, p_max_col, p_max_row))

    hits = []
    for rng in list(ws.merged_cells.ranges):
        # check if it's a protected area; if so, skip it entirely
        r_min_col, r_min_row, r_max_col, r_max_row = (
            rng.min_col,
            rng.min_row,
            rng.max_col,
            rng.max_row,
        )

        is_protected = any(
            (r_min_col == p_min_col and
             r_max_col == p_max_col and
             r_min_row == p_min_row and
             r_max_row == p_max_row)
            for (p_min_col, p_min_row, p_max_col, p_max_row) in protected_bounds
        )
        if is_protected:
            continue

        same_row = (r_min_row <= row <= r_max_row)
        overlap_col = not (r_max_col < min_col or r_min_col > max_col)
        if same_row and overlap_col:
            hits.append(rng)

    # unmerge hits
    for rng in hits:
        ws.unmerge_cells(str(rng))



def write_rows(ws, start_row: int, cols_letters: List[str], rows: List[Tuple[str, float, str]]):
    cols = [col_letter_to_index(x) for x in cols_letters]
    r = start_row
    for name, amount, unit in rows:
        cell_name = ws.cell(row=r, column=cols[0], value=name)
        cell_name.fill = WHITE_FILL
        cell_name.font = BLUE_FONT
        cell_name.alignment = LEFT_ALIGN
        cell_name.border = THICK_LEFT

        cell_amount = ws.cell(row=r, column=cols[1], value=amount)
        cell_amount.number_format = "0.0#####"   # round to 5 decimal places
        cell_amount.fill = WHITE_FILL
        cell_amount.font = BLUE_FONT
        cell_amount.alignment = CENTER_ALIGN
        cell_amount.border = THIN

        cell_unit = ws.cell(row=r, column=cols[2], value=unit)
        cell_unit.fill = WHITE_FILL
        cell_unit.font = BLUE_FONT
        cell_unit.alignment = CENTER_ALIGN
        cell_unit.border = THIN

        r += 1

def generate_report(product_id: int, template_xlsx: str, output_xlsx: str):
    record_json_path = generate_json(product_id)  
    data = load_json(record_json_path)
    wb = openpyxl.load_workbook(template_xlsx)
    stage_rows = collect_by_stage(data)
    product_name = collect_product_name(data)

    ws = wb.active
    ws["C9"].value = product_name

    stage_anchor_rows = {}
    for stage, cfg in TARGET_STAGE_IN_EXCEL.items():
        anchor = cfg["anchor"]          # "B172"
        row_idx = anchor_to_row(anchor) # 172
        stage_anchor_rows[stage] = row_idx
    STAGE_ORDER = ["原料取得", "製造", "配銷", "使用", "廢棄"]

    for stage in STAGE_ORDER:
        rows = stage_rows.get(stage)
        if not rows:
            continue # no data for this stage

        cfg  = TARGET_STAGE_IN_EXCEL.get(stage, {})
        cols = cfg.get("cols")
        if not cols:
            print(f"no cols, skip {stage}")
            continue

        # use the possibly-updated anchor row
        header_row = stage_anchor_rows[stage]
        start_row  = header_row + 1

        data_row, extra, boundary = ensure_space_for_rows(ws, start_row, cols, len(rows))

        # unmerge + copy style
        for i in range(len(rows)):
            unmerge_only_on_row_for_cols(ws, data_row + i, cols)

        # write data
        write_rows(ws, data_row, cols, rows)

        # update subsequent stage anchor rows
        if extra > 0:
            for later_stage in STAGE_ORDER:
                if later_stage == stage:
                    continue
                if stage_anchor_rows[later_stage] >= boundary:
                    stage_anchor_rows[later_stage] += extra

    wb.save(output_xlsx)
    print(f"output file: {output_xlsx}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("template_xlsx")
    ap.add_argument("record_json")
    ap.add_argument("output_xlsx")
    args = ap.parse_args()

    generate_report(args.template_xlsx, args.record_json, args.output_xlsx)


if __name__ == "__main__":
    main()


